from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl.chart import PieChart, Reference

saud = 0
nsaud = 0
nuti = 0

def main_program():
    def get_answers():
        global saud, nsaud, nuti
        try:
            result = int(answers1.get()) + int(answers2.get()) + int(answers3.get())
        except ValueError:
            result = 0

        if result >= 100:
            saud += 1
            categoria = 'saudável'
        elif result >= -30 and result <= 70:
            nsaud += 1
            categoria = 'não saudável'
        else:
            nuti += 1
            categoria = 'meio-termo'

        messagebox.showinfo("Resultado", f"Categoria: {categoria}\nResultado: {result}")
        update_excel()

    window = Tk()
    window.geometry("1920x1080")
    window.configure(background="#FFCC66")
    window.title("Gráfico")

    answers1 = IntVar()
    answers2 = IntVar()
    answers3 = IntVar()

    question1 = Label(window, text="1 - Quanto tempo você passa utilizando dispositivos eletrônicos por dia?", 
                      background="#FFCC66", foreground="#000", font=("Helvetica", 13))
    question1.pack()

    Radiobutton(window, text="a) 2 - 3 horas", background="#FFCC66", value=50, variable=answers1, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="b) Mais de 3 horas", background="#FFCC66", value=10, variable=answers1, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="c) Não utiliza", background="#FFCC66", value=-50, variable=answers1, font=("Helvetica", 13)).pack()

    question2 = Label(window, text="2 - Quanto você fica nas redes sociais?", 
                      background="#FFCC66", foreground="#000", font=("Helvetica", 13))
    question2.pack()

    Radiobutton(window, text="a) 1 - 2 horas", background="#FFCC66", value=50, variable=answers2, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="b) Mais de 3 horas", background="#FFCC66", value=10, variable=answers2, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="c) Não utiliza", background="#FFCC66", value=-50, variable=answers2, font=("Helvetica", 13)).pack()

    question3 = Label(window, text="3 - Quanto tempo você faz pausa?", 
                      background="#FFCC66", foreground="#000", font=("Helvetica", 13))
    question3.pack()

    Radiobutton(window, text="a) 30 - 60 minutos", background="#FFCC66", value=50, variable=answers3, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="b) Raramente faz pausa", background="#FFCC66", value=10, variable=answers3, font=("Helvetica", 13)).pack()
    Radiobutton(window, text="c) Não se aplica", background="#FFCC66", value=-50, variable=answers3, font=("Helvetica", 13)).pack()

    submit_button = Button(window, text="Enviar resposta", command=get_answers)
    submit_button.place(x=912, y=360, width=100, height=20)

    window.mainloop()

def update_excel():
    global saud, nsaud, nuti

    try:
        book = openpyxl.load_workbook("Planilha_do_Grafico.xlsx")
        graphic_page = book["Grafico"]
    except FileNotFoundError:
        book = openpyxl.Workbook()
        graphic_page = book.create_sheet("Grafico", 0)
        graphic_page.append(["Categoria", "Total"])
        book.save("Planilha_do_Grafico.xlsx")

    graphic_page["A2"] = "Saudável"
    graphic_page["B2"] = saud
    graphic_page["A3"] = "Não saudável"
    graphic_page["B3"] = nsaud
    graphic_page["A4"] = "Meio-termo"
    graphic_page["B4"] = nuti

    chart = PieChart()
    chart.title = "Gráfico de Uso Tecnológico"
    data = Reference(graphic_page, min_col=2, min_row=1, max_row=4)
    labels = Reference(graphic_page, min_col=1, min_row=2, max_row=4)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    graphic_page.add_chart(chart, "H1")

    total_pessoas = saud + nsaud + nuti
    info_text = f"Total de pessoas: {total_pessoas}\nSaudáveis: {saud}\nNão saudáveis: {nsaud}\nMeio-termo: {nuti}"
    graphic_page["D5"] = info_text

    book.save("Planilha_do_Grafico.xlsx")

main_program()
